import io
import openpyxl  # type: ignore

def test_import_excel_sin_archivo(client, auth_header_residente):
    response = client.post("/api/grupos/importar-excel", headers=auth_header_residente)
    assert response.status_code == 422

def test_importar_excel_sermul_exitoso(client, auth_header_residente):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GRUPOS"
    ws.cell(row=10, column=1, value="GRUPO 4-MANT. DE CAJON")
    ws.cell(row=11, column=1, value="Nro.")
    ws.cell(row=11, column=2, value="NOMBRES Y APELLIDOS")
    ws.cell(row=11, column=3, value="DNI")
    ws.cell(row=11, column=4, value="POSICION")
    
    ws.cell(row=12, column=1, value=1)
    ws.cell(row=12, column=2, value="CHURA CHURA YUNIER PASCULI")
    ws.cell(row=12, column=3, value="70937958")
    ws.cell(row=12, column=4, value="SUPERVISOR DE OPERACION")
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    files = {"file": ("grupos.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {"parada_id": 1}
    
    response = client.post("/api/grupos/importar-excel", files=files, data=data, headers=auth_header_residente)
    assert response.status_code == 201

def test_import_excel_formato_estandar(client, auth_header_residente):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="GRUPO 5-ESTANDAR")
    ws.cell(row=2, column=1, value="Nro.")
    ws.cell(row=2, column=2, value="NOMBRES Y APELLIDOS")
    ws.cell(row=2, column=3, value="DNI")
    ws.cell(row=2, column=4, value="POSICION")
    
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=2, value="USUARIO NUEVO UNO")
    ws.cell(row=3, column=3, value="88888881")
    ws.cell(row=3, column=4, value="MECANICO")

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    files = {"file": ("grupos.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {"parada_id": 1}
    
    response = client.post("/api/grupos/importar-excel", files=files, data=data, headers=auth_header_residente)
    assert response.status_code == 201
    res = response.json()["resumen"]
    assert res["filas_procesadas"] > 0
    assert res["usuarios_creados"] + res["usuarios_actualizados"] > 0
    assert res["grupos_creados"] + res["grupos_actualizados"] > 0

def test_import_excel_dni_invalido(client, auth_header_residente):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="GRUPO 6-INVALIDO")
    ws.cell(row=2, column=2, value="NOMBRES")
    ws.cell(row=2, column=3, value="DNI")
    ws.cell(row=2, column=4, value="POSICION")
    
    ws.cell(row=3, column=2, value="DNI CORTO")
    ws.cell(row=3, column=3, value="12345") # Invalido
    ws.cell(row=3, column=4, value="MECANICO")

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    response = client.post("/api/grupos/importar-excel", files={"file": ("grupos.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, data={"parada_id": 1}, headers=auth_header_residente)
    assert response.status_code == 201
    assert response.json()["resumen"]["errores"] > 0

def test_import_excel_fila_sin_dni(client, auth_header_residente):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="GRUPO 7-SINDNI")
    ws.cell(row=2, column=2, value="NOMBRES")
    ws.cell(row=2, column=3, value="DNI")
    ws.cell(row=2, column=4, value="POSICION")
    
    ws.cell(row=3, column=2, value="SIN DNI")
    ws.cell(row=3, column=3, value="") # Vacio
    ws.cell(row=3, column=4, value="MECANICO")

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    response = client.post("/api/grupos/importar-excel", files={"file": ("grupos.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, data={"parada_id": 1}, headers=auth_header_residente)
    assert response.status_code == 201
    assert response.json()["resumen"]["filas_procesadas"] == 0

def test_import_excel_grupo_ya_existe(client, auth_header_residente):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="GRUPO 8-DUPLICADO")
    ws.cell(row=2, column=3, value="DNI")
    ws.cell(row=3, column=2, value="DUPLICADO UNO")
    ws.cell(row=3, column=3, value="88888882")
    ws.cell(row=3, column=4, value="MECANICO")

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    client.post("/api/grupos/importar-excel", files={"file": ("grupos.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, data={"parada_id": 1}, headers=auth_header_residente)
    
    file_stream.seek(0)
    response = client.post("/api/grupos/importar-excel", files={"file": ("grupos.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, data={"parada_id": 1}, headers=auth_header_residente)
    assert response.status_code == 201
    res = response.json()["resumen"]
    assert res["grupos_creados"] == 0
    assert res["grupos_actualizados"] > 0

def test_import_excel_reasigna_usuario(client, auth_header_residente):
    wb_a = openpyxl.Workbook()
    ws_a = wb_a.active
    ws_a.cell(row=1, column=1, value="GRUPO 9-A")
    ws_a.cell(row=2, column=3, value="DNI")
    ws_a.cell(row=3, column=2, value="REASIGNADO")
    ws_a.cell(row=3, column=3, value="88888883")
    ws_a.cell(row=3, column=4, value="MECANICO")
    fs_a = io.BytesIO()
    wb_a.save(fs_a)
    fs_a.seek(0)
    client.post("/api/grupos/importar-excel", files={"file": ("g.xlsx", fs_a, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, data={"parada_id": 1}, headers=auth_header_residente)
    
    wb_b = openpyxl.Workbook()
    ws_b = wb_b.active
    ws_b.cell(row=1, column=1, value="GRUPO 9-B")
    ws_b.cell(row=2, column=3, value="DNI")
    ws_b.cell(row=3, column=2, value="REASIGNADO")
    ws_b.cell(row=3, column=3, value="88888883")
    ws_b.cell(row=3, column=4, value="MECANICO")
    fs_b = io.BytesIO()
    wb_b.save(fs_b)
    fs_b.seek(0)
    response = client.post("/api/grupos/importar-excel", files={"file": ("g.xlsx", fs_b, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, data={"parada_id": 1}, headers=auth_header_residente)
    
    assert response.status_code == 201
    res = response.json()["resumen"]
    assert res["usuarios_actualizados"] > 0
